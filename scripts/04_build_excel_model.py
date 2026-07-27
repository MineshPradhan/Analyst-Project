"""
04_build_excel_model.py
------------------------
Builds the client-facing Excel deliverable (per JD: "Developing and
maintaining complex Excel... models"). Live formulas, not hardcoded numbers,
so the client can flex the weighting assumptions and see the ranking update.
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

BASE = os.path.join(os.path.dirname(__file__), "..")
DATA = os.path.join(BASE, "data")
OUT = os.path.join(BASE, "outputs")
os.makedirs(OUT, exist_ok=True)

combined = pd.read_csv(f"{DATA}/combined_clean.csv")

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
BOLD = Font(bold=True)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_NAME = "Arial"

wb = Workbook()

# ============================================================
# SHEET 1: Assumptions (editable inputs - blue/yellow)
# ============================================================
ws1 = wb.active
ws1.title = "Assumptions"
ws1["A1"] = "US State Expansion — Model Assumptions"
ws1["A1"].font = Font(bold=True, size=14, name=FONT_NAME)
ws1.merge_cells("A1:C1")

labels = [
    ("Weight: Market Size (population)", 0.20),
    ("Weight: Population Growth", 0.20),
    ("Weight: Median Income", 0.15),
    ("Weight: Low Competition", 0.20),
    ("Weight: Low Cost (rent+buildout)", 0.15),
    ("Weight: Low Unemployment", 0.10),
    ("Assumed Operating Margin", 0.14),
]
ws1["A3"] = "Assumption"
ws1["B3"] = "Value"
for cell in ["A3", "B3"]:
    ws1[cell].font = HEADER_FONT
    ws1[cell].fill = HEADER_FILL
row = 4
for label, val in labels:
    ws1[f"A{row}"] = label
    ws1[f"A{row}"].font = Font(name=FONT_NAME)
    c = ws1[f"B{row}"]
    c.value = val
    c.font = BLUE
    c.fill = YELLOW_FILL
    c.number_format = "0.0%"
    row += 1
ws1["A12"] = "Note: weights above are the editable levers. Change B4:B9 and the"
ws1["A13"] = "Ranked Model sheet recalculates the opportunity score automatically."
ws1["A12"].font = Font(italic=True, size=9, name=FONT_NAME)
ws1["A13"].font = Font(italic=True, size=9, name=FONT_NAME)
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 14

# ============================================================
# SHEET 2: Raw Combined Data
# ============================================================
ws2 = wb.create_sheet("Raw Data")
headers = list(combined.columns)
for j, h in enumerate(headers, start=1):
    c = ws2.cell(row=1, column=j, value=h.replace("_", " ").title())
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
for i, r in combined.iterrows():
    for j, h in enumerate(headers, start=1):
        c = ws2.cell(row=i + 2, column=j, value=r[h])
        c.font = Font(name=FONT_NAME)
        c.border = BORDER
for j, h in enumerate(headers, start=1):
    ws2.column_dimensions[get_column_letter(j)].width = max(14, len(h) + 2)
n_states = len(combined)

# ============================================================
# SHEET 3: Ranked Model (formulas reference Raw Data + Assumptions)
# ============================================================
ws3 = wb.create_sheet("Ranked Model")
cols = ["State", "Population (mm)", "Pop Growth %", "Median Income", "Unemployment %",
        "Competitor Stores", "Rent/SqFt", "Buildout Cost",
        "Score: Size", "Score: Growth", "Score: Income", "Score: Low Comp",
        "Score: Low Cost", "Score: Low Unemp", "Opportunity Score",
        "Predicted Revenue/Store", "Est. Annual Profit/Store", "Payback (Years)",
        "Cost Index (helper)"]
for j, h in enumerate(cols, start=1):
    c = ws3.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

# Map Raw Data columns -> letters
rd_col = {h: get_column_letter(idx + 1) for idx, h in enumerate(headers)}

# Predicted revenue: simplified linear proxy so the SHEET is self-contained
# (uses same signal direction as the Python regression: bigger, growing,
# richer, less-crowded markets predict higher revenue per store)
for i in range(n_states):
    r = i + 2
    rd_r = i + 2
    ws3.cell(row=r, column=1, value=f"='Raw Data'!{rd_col['state']}{rd_r}")
    ws3.cell(row=r, column=2, value=f"='Raw Data'!{rd_col['population_2025_millions']}{rd_r}")
    ws3.cell(row=r, column=3, value=f"='Raw Data'!{rd_col['population_growth_5yr_pct']}{rd_r}")
    ws3.cell(row=r, column=4, value=f"='Raw Data'!{rd_col['median_household_income']}{rd_r}")
    ws3.cell(row=r, column=5, value=f"='Raw Data'!{rd_col['unemployment_rate_pct']}{rd_r}")
    ws3.cell(row=r, column=6, value=f"='Raw Data'!{rd_col['competitor_store_count']}{rd_r}")
    ws3.cell(row=r, column=7, value=f"='Raw Data'!{rd_col['avg_commercial_rent_per_sqft']}{rd_r}")
    ws3.cell(row=r, column=8, value=f"='Raw Data'!{rd_col['avg_buildout_cost']}{rd_r}")

    # Normalized scores (min-max across the 20-state range) — columns B,C,D,F,(G+H),E
    rng = lambda col: f"B$2:B${n_states+1}" if col == "B" else None
    def norm(col_letter, invert=False):
        rng_ = f"{col_letter}$2:{col_letter}${n_states+1}"
        if not invert:
            return f"({col_letter}{r}-MIN({rng_}))/(MAX({rng_})-MIN({rng_}))"
        else:
            return f"1-({col_letter}{r}-MIN({rng_}))/(MAX({rng_})-MIN({rng_}))"

    ws3.cell(row=r, column=9, value=f"={norm('B')}")                     # size
    ws3.cell(row=r, column=10, value=f"={norm('C')}")                    # growth
    ws3.cell(row=r, column=11, value=f"={norm('D')}")                    # income
    ws3.cell(row=r, column=12, value=f"={norm('F', invert=True)}")       # low competition
    ws3.cell(row=r, column=19, value=f"=G{r}+H{r}/1000")                 # cost helper (plain col)
    ws3.cell(row=r, column=13, value=f"={norm('S', invert=True)}")       # low cost
    ws3.cell(row=r, column=14, value=f"={norm('E', invert=True)}")       # low unemployment

    ws3.cell(row=r, column=15,
              value=(f"=(I{r}*Assumptions!$B$4+J{r}*Assumptions!$B$5+K{r}*Assumptions!$B$6+"
                      f"L{r}*Assumptions!$B$7+M{r}*Assumptions!$B$8+N{r}*Assumptions!$B$9)*100"))

    # Predicted revenue proxy formula (transparent, in-sheet regression proxy)
    ws3.cell(row=r, column=16,
              value=f"=800000+(B{r}*15000)+(C{r}*8000)+(D{r}*6)-(F{r}*900)")
    ws3.cell(row=r, column=17, value=f"=P{r}*Assumptions!$B$10")
    ws3.cell(row=r, column=18, value=f"=H{r}/Q{r}")

    for col in range(1, 20):
        ws3.cell(row=r, column=col).border = BORDER
        ws3.cell(row=r, column=col).font = Font(name=FONT_NAME)

# number formats
for r in range(2, n_states + 2):
    ws3.cell(row=r, column=2).number_format = "0.00"
    ws3.cell(row=r, column=3).number_format = "0.0%"
    for col in (4, 16, 17, 18):
        pass
    ws3.cell(row=r, column=4).number_format = "$#,##0"
    ws3.cell(row=r, column=5).number_format = "0.0%"
    for col in [9, 10, 11, 12, 13, 14]:
        ws3.cell(row=r, column=col).number_format = "0.00"
    ws3.cell(row=r, column=15).number_format = "0.0"
    ws3.cell(row=r, column=16).number_format = "$#,##0"
    ws3.cell(row=r, column=17).number_format = "$#,##0"
    ws3.cell(row=r, column=18).number_format = "0.00\"y\""

for j in range(1, 20):
    ws3.column_dimensions[get_column_letter(j)].width = 16
ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["S"].width = 10

# Conditional-ish highlight: top score bold (static highlight applied post-recalc not needed; skip)

# Chart: Opportunity Score by state
chart = BarChart()
chart.title = "Opportunity Score by State"
chart.y_axis.title = "Score (0-100)"
chart.x_axis.title = "State"
data_ref = Reference(ws3, min_col=15, min_row=1, max_row=n_states + 1)
cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=n_states + 1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 24
chart.height = 12
ws3.add_chart(chart, "T2")

wb.save(f"{OUT}/US_Expansion_Client_Model.xlsx")
print("Saved:", f"{OUT}/US_Expansion_Client_Model.xlsx")
